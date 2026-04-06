from tkinter import *
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from datetime import datetime

file_name = "PingPongData.xlsx"

# --- EXCEL SETUP ---
try:
    wb = load_workbook(file_name)
except:
    wb = Workbook()

if "Matches" not in wb.sheetnames:
    sheet = wb.active
    sheet.title = "Matches"
    sheet.append(["Date", "Player 1", "Player 2", "Score", "Winner", "Game To", "Faults P1", "Faults P2"])
else:
    sheet = wb["Matches"]

# --- GAME VARIABLES ---
score1 = 0
score2 = 0
fault1 = 0        # temporary fault counter toward 2 faults
fault2 = 0
total_faults1 = 0 # cumulative faults for stats
total_faults2 = 0
game_to = 15

# --- FUNCTIONS ---
def start_game():
    global score1, score2, fault1, fault2, game_to
    p1 = player1_var.get()
    p2 = player2_var.get()
    if not p1 or not p2:
        status_label.config(text="❌ Select both players")
        return

    score1 = score2 = fault1 = fault2 = 0
    try:
        game_to = int(game_to_entry.get())
    except:
        game_to = 15

    setup_frame.pack_forget()
    game_frame.pack(expand=True, fill=BOTH)

    header_label.config(text="🏓 PING PONG SCOREBOARD 🏓")
    player1_label.config(text=p1)
    player2_label.config(text=p2)

    winner_label.config(text="")
    update_display()
    status_label.config(text="")

def update_display():
    score1_label.config(text=str(score1))
    score2_label.config(text=str(score2))
    fault1_label.config(text=f"FAULT: {fault1}")
    fault2_label.config(text=f"FAULT: {fault2}")

def add_point(player):
    global score1, score2, fault1, fault2
    if player == 1:
        score1 += 1
        fault1 = 0
        fault2 = 0
    else:
        score2 += 1
        fault1 = 0
        fault2 = 0
    update_display()
    check_winner()

def subtract_point(player):
    global score1, score2
    if player == 1 and score1 > 0:
        score1 -= 1
    elif player == 2 and score2 > 0:
        score2 -= 1
    update_display()

def add_fault(player):
    global fault1, fault2, score1, score2, total_faults1, total_faults2
    if player == 1:
        fault1 += 1
        total_faults1 += 1
        if fault1 >= 2:
            score2 += 1
            fault1 = 0
            fault2 = 0
    else:
        fault2 += 1
        total_faults2 += 1
        if fault2 >= 2:
            score1 += 1
            fault2 = 0
            fault1 = 0
    update_display()
    check_winner()

def subtract_fault(player):
    global fault1, fault2, total_faults1, total_faults2
    if player == 1 and fault1 > 0 and total_faults1 > 0:
        fault1 -= 1
        total_faults1 -= 1
    elif player == 2 and fault2 > 0 and total_faults2 > 0:
        fault2 -= 1
        total_faults2 -= 1
    update_display()

def check_winner():
    # Win only if someone reached game_to AND leads by 2
    if (score1 >= game_to or score2 >= game_to) and abs(score1 - score2) >= 2:
        winner = player1_var.get() if score1 > score2 else player2_var.get()
        winner_label.config(text=f"🏆 {winner} WINS! 🏆")
        save_match(winner)

def save_match(winner):
    sheet.append([
        datetime.now().strftime("%Y-%m-%d"),
        player1_var.get(),
        player2_var.get(),
        f"{score1}-{score2}",
        winner,
        game_to,
        total_faults1,
        total_faults2
    ])
    wb.save(file_name)

def reset_game():
    game_frame.pack_forget()
    setup_frame.pack(expand=True)

# --- UI SETUP ---
root = Tk()
root.title("🏓 Ping Pong Scoreboard 🏓")
root.configure(bg="#2F2F2F")
root.geometry("900x600")
root.minsize(800, 500)

font_big = ("Courier", 60, "bold")
font_med = ("Courier", 18, "bold")
font_small = ("Courier", 14, "bold")

# --- SETUP FRAME ---
setup_frame = Frame(root, bg="#2F2F2F")
setup_frame.pack(expand=True)

Label(setup_frame, text="🏓 PING PONG SCOREBOARD 🏓", fg="lime", bg="#2F2F2F", font=("Courier", 36, "bold")).pack(pady=20)

player1_var = StringVar()
player2_var = StringVar()

Label(setup_frame, text="Player 1:", fg="white", bg="#2F2F2F", font=font_med).pack(pady=5)
player1_menu = ttk.Combobox(setup_frame, textvariable=player1_var, font=font_med)
player1_menu.pack(pady=5)

Label(setup_frame, text="Player 2:", fg="white", bg="#2F2F2F", font=font_med).pack(pady=5)
player2_menu = ttk.Combobox(setup_frame, textvariable=player2_var, font=font_med)
player2_menu.pack(pady=5)

Label(setup_frame, text="Play To:", fg="white", bg="#2F2F2F", font=font_med).pack(pady=10)
game_to_entry = Entry(setup_frame, font=font_med, justify='center')
game_to_entry.insert(0, "15")
game_to_entry.pack(pady=5)

Button(setup_frame, text="START GAME", command=start_game, bg="darkgray", font=font_med).pack(pady=20)

# --- GAME FRAME ---
game_frame = Frame(root, bg="#2F2F2F")
game_frame.pack_propagate(False)

# --- Header ---
header_label = Label(game_frame, text="", fg="lime", bg="#2F2F2F", font=("Courier", 36, "bold"))
header_label.pack(pady=20)

# --- Main scoreboard frame ---
scoreboard_frame = Frame(game_frame, bg="#2F2F2F")
scoreboard_frame.pack(expand=True, fill=BOTH, padx=50)

# --- Player 1 Column ---
player1_col = Frame(scoreboard_frame, bg="#2F2F2F")
player1_col.pack(side=LEFT, expand=True)

player1_label = Label(player1_col, text="", fg="white", bg="#2F2F2F", font=font_med)
player1_label.pack(pady=(10,5))

score1_label = Label(player1_col, text="0", fg="white", bg="black", font=font_big, width=3)
score1_label.pack(pady=5)

fault1_label = Label(player1_col, text="FAULT: 0", fg="red", bg="#2F2F2F", font=font_med)
fault1_label.pack(pady=5)

# Points Buttons
points_frame1 = Frame(player1_col, bg="#2F2F2F")
points_frame1.pack(pady=5)
Label(points_frame1, text="Point:", fg="white", bg="#2F2F2F", font=font_small).pack(side=LEFT)
Button(points_frame1, text="+1", command=lambda: add_point(1), font=font_small, bg="gray").pack(side=LEFT, padx=5)
Button(points_frame1, text="-1", command=lambda: subtract_point(1), font=font_small, bg="darkgray").pack(side=LEFT, padx=5)

# Fault Buttons (+1 and -1)
faults_frame1 = Frame(player1_col, bg="#2F2F2F")
faults_frame1.pack(pady=5)
Label(faults_frame1, text="Fault:", fg="white", bg="#2F2F2F", font=font_small).pack(side=LEFT)
Button(faults_frame1, text="+1", command=lambda: add_fault(1), font=font_small, bg="red", fg="white").pack(side=LEFT, padx=5)
Button(faults_frame1, text="-1", command=lambda: subtract_fault(1), font=font_small, bg="darkred", fg="white").pack(side=LEFT, padx=5)

# --- Winner label (center) ---
winner_label = Label(scoreboard_frame, text="", fg="yellow", bg="#2F2F2F", font=("Courier", 24, "bold"))
winner_label.pack(side=LEFT, expand=True)

# --- Player 2 Column ---
player2_col = Frame(scoreboard_frame, bg="#2F2F2F")
player2_col.pack(side=RIGHT, expand=True)

player2_label = Label(player2_col, text="", fg="white", bg="#2F2F2F", font=font_med)
player2_label.pack(pady=(10,5))

score2_label = Label(player2_col, text="0", fg="white", bg="black", font=font_big, width=3)
score2_label.pack(pady=5)

fault2_label = Label(player2_col, text="FAULT: 0", fg="red", bg="#2F2F2F", font=font_med)
fault2_label.pack(pady=5)

# Points Buttons
points_frame2 = Frame(player2_col, bg="#2F2F2F")
points_frame2.pack(pady=5)
Label(points_frame2, text="Point:", fg="white", bg="#2F2F2F", font=font_small).pack(side=LEFT)
Button(points_frame2, text="+1", command=lambda: add_point(2), font=font_small, bg="gray").pack(side=LEFT, padx=5)
Button(points_frame2, text="-1", command=lambda: subtract_point(2), font=font_small, bg="darkgray").pack(side=LEFT, padx=5)

# Fault Buttons (+1 and -1)
faults_frame2 = Frame(player2_col, bg="#2F2F2F")
faults_frame2.pack(pady=5)
Label(faults_frame2, text="Fault:", fg="white", bg="#2F2F2F", font=font_small).pack(side=LEFT)
Button(faults_frame2, text="+1", command=lambda: add_fault(2), font=font_small, bg="red", fg="white").pack(side=LEFT, padx=5)
Button(faults_frame2, text="-1", command=lambda: subtract_fault(2), font=font_small, bg="darkred", fg="white").pack(side=LEFT, padx=5)

# --- New Game Button (bottom center) ---
Button(game_frame, text="NEW GAME", command=reset_game, font=font_med, bg="orange").pack(side=BOTTOM, pady=20)

# --- Dummy players ---
players = ["Alex", "Craig", "Gannon"]
player1_menu['values'] = players
player2_menu['values'] = players

root.mainloop()